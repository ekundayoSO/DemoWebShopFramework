import openpyxl


class RegistrationData:

    test_Registration_data = [
        {"firstname": "Michael", "lastname": "Brown", "email": "michael.mb@email.com", "Password": "tester123",
         "ConfirmPassword": "tester123"},
        {"firstname": "Emily", "lastname": "Johnson", "email": "emily.ej@email.com", "Password": "tester123",
         "ConfirmPassword": "tester123"}]

    @staticmethod
    def getTestData(test_case_name):
        Dict = {}

        # Read from the Excel sheet
        book = openpyxl.load_workbook("C:\\Users\\omowu\\Desktop\\Python\\data.xlsx")
        sheet = book.active

        for i in range(1, sheet.max_row + 1):
            if sheet.cell(row=i, column=1).value == test_case_name:
                for j in range(2, sheet.max_column + 1):
                    Dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value
        return [Dict]
