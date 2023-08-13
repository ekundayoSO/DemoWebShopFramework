import openpyxl

# Read from the Excel sheet
book = openpyxl.load_workbook("C:\\Users\\omowu\\Desktop\\Python\\data.xlsx")
sheet = book.active


cell = sheet.cell(row=1, column=2)
print(cell.value)

# Write to the Excel sheet
sheet.cell(row=2, column=2).value = "Sulaimon"
print(sheet.cell(row=2, column=2).value)

# Getting maximum number of rows and columns in a sheet
print(sheet.max_row)
print(sheet.max_column)

print(sheet['B6'].value)


